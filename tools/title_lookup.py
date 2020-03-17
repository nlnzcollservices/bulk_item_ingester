from openpyxl import Workbook, load_workbook
from tools.title_ref_updater import update_titles
import os
import shutil
from datetime import datetime

my_file = "titles_reference.xlsx"

shared_root = r"G:\Fileplan\Bib_Services\Non-Clio_formats\Acquisitions Team\bulk item ingest"

def get_xlsx_spreadsheet(f):
	titles_lookup = {"PROD":{}, "SAND":{}}
	len_data = 11
	my_data = []
	wb = load_workbook(f, data_only=True)
	ws = wb.active
	row_counter = 0
	cell_counter = 0
	for i, row in enumerate(ws):
		my_row = []
		row_counter  += 1
		if row_counter == 1:
			pass
		else:
			for cell in row[0:len_data]:
				try:
					my_row.append(int(cell.value))
				except:
					my_row.append(cell.value)

			if my_row != [None for x in range(len_data)]:
				my_data.append(my_row[0:len_data])

	for row in my_data:
		if row != [None, None, None, None, None, None, None]:
			### these are all gubby excel fixes. :( 
			mms = row[1]
			if str(mms).endswith("0"):
				mms+=6
			atl = row[4]
			if str(atl).endswith("00"):
				atl += 36
			wn  = row[3]
			if str(wn).endswith("00"):
				wn += 36
			if row[5]:
				my_system = row[5]
			else:
				my_system = 'PROD'
			titles_lookup[my_system][mms] = {"mms":mms, "ATL":atl, "WN":wn, "pol":row[2], "title":row[0], 'system':my_system, 'signed_off':row[6]}

	return titles_lookup

def add_new_title_to_spreadsheet(new_item_dict):
	in_existing = False
	my_ref_workbook = "titles_reference.xlsx"
	all_mms_ids = {}
	if new_item_dict['title'].endswith("."):
		new_item_dict['title'] = new_item_dict['title'][:-1]

	old_data = get_xlsx_spreadsheet(my_ref_workbook)

	for system, titles in old_data.items():
		for mms in titles:
			if mms not in all_mms_ids: 
				all_mms_ids[mms] = []
			all_mms_ids[mms].append(old_data[system][mms]['system'])


	
	if new_item_dict['mms'] in all_mms_ids and new_item_dict['system'] in all_mms_ids[new_item_dict['mms']] :
		in_existing = True

	if not in_existing:
		wb = load_workbook(my_ref_workbook)
		ws = wb.active


		# gets rid of empty rows
		empty_rows = []
		for idx, row in enumerate(ws.iter_rows(max_col=50), start=1):
			empty = not any((cell.value for cell in row))
			if empty:
				empty_rows.append(idx)
		for row_idx in reversed(empty_rows):
			ws.delete_rows(row_idx, 1)

		max_row_id = ws.max_row

		if new_item_dict['system'] == "SAND":
			for cell_id, entry in enumerate( [new_item_dict['title'], new_item_dict['mms'], new_item_dict['pol'], new_item_dict['WN'], new_item_dict['ATL'], new_item_dict['system'], "Y"], start=1):
				ws.cell(row=max_row_id+1, column=cell_id, value=entry)
		elif new_item_dict['system'] == "PROD":
			for cell_id, entry in enumerate( [new_item_dict['title'], new_item_dict['mms'], new_item_dict['pol'], new_item_dict['WN'], new_item_dict['ATL'], new_item_dict['system'], ""], start=1):
				ws.cell(row=max_row_id+1, column=cell_id, value=entry)


		new_local_name = os.path.join("titles_reference_backups", my_ref_workbook.replace(".xlsx", f"_{datetime.now().strftime('%Y-%m-%d %H_%M_%S')}.xlsx"))
		if os.path.exists(new_local_name):
			os.remove(new_local_name)
		os.rename("titles_reference.xlsx", new_local_name)
		wb.save(my_ref_workbook)
		wb.save(os.path.join( shared_root, my_ref_workbook))

		print (f"{new_item_dict['title']} - {new_item_dict['system']} added to lookup")
	
	else:
		print (f"{new_item_dict['title']} - {new_item_dict['system']} is a dup mms id. Not added")


titles_lookup = get_xlsx_spreadsheet(my_file)